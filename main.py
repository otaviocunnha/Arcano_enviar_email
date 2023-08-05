from PySide6.QtWidgets import (QApplication, QMainWindow)
from ui_main import Ui_MainWindow
from PySide6.QtGui import QIcon
import sys
import win32com.client as win32
import os
import pandas as pd
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import *
from openpyxl import load_workbook


class MainWindow(QMainWindow, Ui_MainWindow):
    
    def __init__(self):
        
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("Arcano")
        appIcon = QIcon(u"icon.ico")
        self.setWindowIcon(appIcon)

        
        # Define o ícone do aplicativo na barra de tarefas
        self.trayIcon = QSystemTrayIcon(self)
        self.trayIcon.setIcon(appIcon)
        self.trayIcon.show()
        
        
        # Botão para enviar o e-mail
        self.btnExecute.clicked.connect(self.send_email)
          
        #  Botão para editar 
        self.btn_consulta.clicked.connect(self.consultar_unidade)
    
        # Botão para enviar o e-mail outros
        self.btn_send_email.clicked.connect(self.send_email_outros)
        
        
        
    def send_email(self):
        
        # N fatura Aereo 
        numero_aereo = self.lineEdit.text()
        
        # N Fatura hotel 
        numero_hotel = self.lineEdit_2.text()
        
        # caminho que a pessoa vai colocar o excel da planilha aereo 
        excel_aereo_planilha = self.planilha_1.text()
        
        # caminho que a pessoa vai colocar o excel da planilha hotel
        excel_hotel_planilha = self.planilha_2.text()
        
        # caminho que a pessoa vai colocar o excel da planilha fatura 
        excel_fatura_planilha = self.planilha_3.text()    
        
        # tirar informações que podem vir na hora do usuario colar o caminho  
        excel_aereo_planilha = pd.read_excel(f'{excel_aereo_planilha}'.replace('"',''), header=7)
        
        # tirar informações que podem vir na hora do usuario colar o caminho  
        excel_hotel = pd.read_excel(f'{excel_hotel_planilha}'.replace('"',''), header=7)


        if excel_fatura_planilha:    
            # tirar informações que podem vir na hora do usuario colar o caminho  
            excel_fatura = pd.read_excel(f'{excel_fatura_planilha}'.replace('"',''))
        else: 
            excel_fatura = None
        
        # Vai pegar somente as colunas que eu quero     
        colunas_aereo = ["Centro de custo", "Nome das colunas"]
        
        # Vai pegar somente as colunas que eu quero 
        colunas_hotel = ["Centro de custo","Nome das colunas" ]

        # Vai pegar somente as colunas que eu quero 
        coluna_fatura = ["Centro de custo", "Nome das colunas"]


        # Selecionar apenas as colunas desejadas para cada arquivo
        dados_aereo = excel_aereo_planilha[colunas_aereo]
        # Selecionar apenas as colunas desejadas para cada arquivo
        dados_hotel = excel_hotel[colunas_hotel]
        

        if excel_fatura is not None:        
            # Selecionar apenas as colunas desejadas para cada arquivo   
            dados_fatura = excel_fatura[coluna_fatura]


        # Vai criar um excel novo com as colunas que selecionei anteriormente     
        with pd.ExcelWriter('dados_selecionados.xlsx', engine='openpyxl') as writer:
            dados_aereo.to_excel(writer, sheet_name='aereo', index=False)
            dados_hotel.to_excel(writer, sheet_name='hotel', index=False)
            if excel_fatura is not None:
                dados_fatura.to_excel(writer, sheet_name='fatura', index=False)
        

        # Ler a planilha 'aereo' no arquivo Excel de cada aba com o excel que criei 
        banco_dados = pd.read_excel("dados_selecionados.xlsx", sheet_name='aereo')
        
        # Ler a planilha 'hotel' no arquivo Excel de cada aba com o excel que criei 
        banco_dados_hotel = pd.read_excel("dados_selecionados.xlsx", sheet_name='hotel')
        
        if excel_fatura is not None: 
            # Ler a planilha 'fatura' no arquivo Excel de cada aba com o excel que criei 
            banco_dados_fatura = pd.read_excel("dados_selecionados.xlsx", sheet_name='fatura')
    

        # carregar o arquivo Excel existente
        workbook = load_workbook("dados_selecionados.xlsx")

        #Coloca informações no campo de atualização
        self.atualizacao.setText('STATUS')

        # agrupar os dados por 'Centro de custo nome'
        grupos = banco_dados.groupby('Centro de custo nome')
        
        grupos_hotel = banco_dados_hotel.groupby('Centro de custo nome')

        if excel_fatura is not None:   
            grupos_fatura = banco_dados_fatura.groupby('Centro de custo nome')
        
        
        def send_mail_aereo():    
            for nome_loja, grupo in grupos:
                # selecionar os dados relevantes para cada loja
                dados_loja = grupo
                # Ele ve quais os centros de de custos iguais e cria uma nova aba chamada com o centro de custo e coloca 'Aereo' no nome na aba AEREO
                if not dados_loja.empty:
                    worksheet = workbook.create_sheet(f"{nome_loja}_aereo")
                    worksheet.append(["Nome das colunas"])
                    for index, row in dados_loja.iterrows():
                        worksheet.append(row.tolist())
                    workbook.save(filename='dados_selecionados.xlsx')
                else:
                    self.atualizacao.append(f'Sem gastos aereos na unidade {nome_loja}')
                
                #Conecta com o Outlook
                outlook = win32.Dispatch('outlook.application')

                # Cria um novo e-mail
                mail = outlook.CreateItem(0)

                # ler o excel 
                excel_path = 'dados_selecionados.xlsx'

                # abas que quero com o nome das lojas criadas no código acima
                aba_desejada = [f"{nome_loja}_aereo"]
                
                
                # Define o assunto
                mail.Subject = f'Custo de despesas com gastos aereos da unidade de {aba_desejada[0]}, numero da fatura {numero_aereo}'.replace("_aereo", "")
                
                # Define o corpo da mensagem
                mail.Body = '[NAO RESPONDER ESSE EMAIL] \nOlá, encaminho em anexo os gastos de viagens lançados em seu centro de custo.'
                
                df_dict = {}
                
                #Se tiver as abas selecionadas
                for sheet_name in aba_desejada:
                    try:
                        df_dict[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)
                    except:
                        print(f"A aba {sheet_name} não existe no arquivo Excel") 
                        
                # salva as abas em um novo arquivo
                output_path = f'{nome_loja}.xlsx'

                # le o arquivo que ele criou e a aba
                with pd.ExcelWriter(output_path) as writer:
                    for sheet_name, df in df_dict.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Anexa o arquivo Excel temporário (que contém apenas a aba correspondente)
                attachment = os.path.abspath(output_path)
                mail.Attachments.Add(attachment)
                
                excel = pd.read_excel(f'{nome_loja}.xlsx')
                    
                
                '''Vai comparar se a coluna Centro de custo é igual ao nome da unidade para enviar o e-mail de acordo com cada
                gestor que precisa receber ''' 
                
                if excel['unidade'][0] == 'comparacao':
                    
                    with open('Nome_txt.txt', 'r') as f:
                        email_varzea_grande = f.readlines()
                        email_varzea_grande = [line.strip() for line in email_varzea_grande]
                    mail.To = f'{email_varzea_grande}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                    mail.Send()
                
                elif excel['unidade'][0] == 'comparacao':
                    
                    with open('Nome_txt.txt', 'r') as f:
                        email_Goiainia_goias = f.readlines()
                        email_Goiainia_goias = [line.strip() for line in email_Goiainia_goias]
                    mail.To = f'{email_Goiainia_goias}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                    mail.Send()
                
                elif excel['unidade'][0] == 'comparacao':
                    
                    with open('Nome_txt.txt', 'r') as f:
                        Nome_txt = f.readlines()
                        Nome_txt.txt = [line.strip() for line in Nome_txt]
                    mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")                
                    mail.Send()     
                    
                    
                else:  
                    if nome_loja == 'comparacao':
                        pass 
                    else:
                        self.atualizacao.append(f"A unidade {nome_loja}_aereo não possui cadastro no banco de dados")
    
                    
                
                #Remove o arquivo temporario que criei
                os.remove(output_path) 
              
        def send_mail_hotel():      
            for nome_loja, grupo_hotel in grupos_hotel:
                # selecionar os dados relevantes para cada loja
                dados_loja_hotel = grupo_hotel              
                # Ele ve quais os centros de de custos iguais e cria uma nova aba chamada com o centro de custo e coloca 'Hotel' no nome na aba HOTEL
            
                if not dados_loja_hotel.empty:
                    worksheet = workbook.create_sheet(f"{nome_loja}_hotel")
                    worksheet.append(["Nome das colunas"])
                    for index, row in dados_loja_hotel.iterrows():
                        worksheet.append(row.tolist())
                    workbook.save(filename='dados_selecionados.xlsx')
                else: 
                    self.atualizacao.append(f'Sem gastos hoteleiros na unidade {nome_loja}')
                  
                #Conecta com o Outlook
                outlook = win32.Dispatch('outlook.application')

                # Cria um novo e-mail
                mail = outlook.CreateItem(0)


                # ler o excel 
                excel_path = 'dados_selecionados.xlsx'

                # abas que quero com o nome das lojas criadas no código acima
                aba_desejada = [f"{nome_loja}_hotel"]
                
                
                # Define o assunto
                mail.Subject = f'Custo de despesas com gastos hoteleiros na unidade de {aba_desejada[0]}, numero da fatura {numero_hotel}'.replace("_hotel", "")
                
                # Define o corpo da mensagem
                mail.Body = '[NAO RESPONDER ESSE EMAIL] \nOlá, encaminho em anexo os gastos de viagens lançados em seu centro de custo.' 
                
                
                df_dict = {}
                
                #Se tiver as abas selecionadas
                for sheet_name in aba_desejada:
                    try:
                        df_dict[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)
                    except:
                        print(f"A aba {sheet_name} não existe no arquivo Excel")
                        
                # salva as abas em um novo arquivo
                output_path = f'{nome_loja}.xlsx'

                # le o arquivo que ele criou e a aba
                with pd.ExcelWriter(output_path) as writer:
                    for sheet_name, df in df_dict.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Anexa o arquivo Excel temporário (que contém apenas a aba correspondente)
                attachment = os.path.abspath(output_path)
                mail.Attachments.Add(attachment)
                
                excel = pd.read_excel(f'{nome_loja}.xlsx')
                    
                
                '''Vai comparar se a coluna Centro de custo é igual ao nome da unidade para enviar o e-mail de acordo com cada
                gestor que precisa receber ''' 
                if excel['unidade'][0] == 'comparacao':
                    
                    with open('Nome_txt.txt', 'r') as f:
                        Nome_txt = f.readlines()
                        Nome_txt = [line.strip() for line in Nome_txt]
                    mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                    mail.Send()
                
                elif excel['unidade'][0] == 'comparacao':
                    
                    with open('Nome_txt.txt', 'r') as f:
                        Nome_txt = f.readlines()
                        Nome_txt = [line.strip() for line in Nome_txt]
                    mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                    mail.Send()
                
                elif excel['unidade'][0] == 'comparacao':
                    
                    with open('Nome_txt.txt', 'r') as f:
                        Nome_txt = f.readlines()
                        Nome_txt = [line.strip() for line in Nome_txt]
                    mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")                
                    mail.Send()             
                    
                else:
                    if nome_loja == 'comparacao':
                        continue
                    else:   
                        self.atualizacao.append(f"A unidade {nome_loja}_hotel não possui cadastro no banco de dados")
    
                
                #Remove o arquivo temporario que criei
                os.remove(output_path)

        def send_mail_fatura():  
            
            if excel_fatura is None:      
                self.atualizacao.append(f"A planilha fatura não foi anexada!")

            else:          
                                  
                for nome_loja, grupo_fatura in grupos_fatura:
                    # selecionar os dados relevantes para cada loja
                    if excel_fatura is not None: 
                        dados_loja_fatura = grupo_fatura
                                    
                    # Ele ve quais os centros de de custos iguais e cria uma nova aba chamada com o centro de custo e coloca 'Fatura' no nome na aba FATURA
                
                    if excel_fatura is not None: 
                        if not dados_loja_fatura.empty:
                            worksheet = workbook.create_sheet(f"{nome_loja}_fatura")
                            worksheet.append(["Nome das Colunas"])
                            for index, row in dados_loja_fatura.iterrows():
                                worksheet.append(row.tolist())
                            workbook.save(filename='dados_selecionados.xlsx') 
                        else: 
                            self.atualizacao.append(f'Sem gastos na fatura da unidade {nome_loja}')
                    #Conecta com o Outlook
                    outlook = win32.Dispatch('outlook.application')

                    # Cria um novo e-mail
                    mail = outlook.CreateItem(0)


                    # ler o excel 
                    excel_path = 'dados_selecionados.xlsx'

                    # abas que quero com o nome das lojas criadas no código acima
                    aba_desejada = [f"{nome_loja}_fatura"]
                    
                    
                    # Define o assunto
                    mail.Subject = f'Custo de despesas com gastos da fatura na unidade de {aba_desejada[0]}'.replace("_fatura", "")
                    
                    # Define o corpo da mensagem
                    mail.Body = '[NAO RESPONDER ESSE EMAIL] \nOlá, encaminho em anexo os gastos de viagens lançados em seu centro de custo. '
                
                    
                    df_dict = {}
                    
                    #Se tiver as abas selecionadas
                    for sheet_name in aba_desejada:
                        try:
                            df_dict[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)
                        except:
                            print(f"A aba {sheet_name} não existe no arquivo Excel") #TODO: Olhar esse print 
                            
                    # salva as abas em um novo arquivo
                    output_path = f'{nome_loja}.xlsx'

                    # le o arquivo que ele criou e a aba
                    with pd.ExcelWriter(output_path) as writer:
                        for sheet_name, df in df_dict.items():
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Anexa o arquivo Excel temporário (que contém apenas a aba correspondente)
                    attachment = os.path.abspath(output_path)
                    mail.Attachments.Add(attachment)
                    
                    excel = pd.read_excel(f'{nome_loja}.xlsx')
                        
                    
                    '''Vai comparar se a coluna Centro de custo é igual ao nome da unidade para enviar o e-mail de acordo com cada
                    gestor que precisa receber ''' 
                    if excel['unidade'][0] == 'comparação':
                        
                        with open('Nome_txt.txt', 'r') as f:
                            Nome_txt = f.readlines()
                            Nome_txt = [line.strip() for line in Nome_txt]
                        mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                        mail.Send()
                    
                    elif excel['unidade'][0] == 'comparação':
                        
                        with open('email_goiania_goias.txt', 'r') as f:
                            Nome_txt = f.readlines()
                            Nome_txt = [line.strip() for line in Nome_txt]
                        mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                        mail.Send()
                    
                    elif excel['unidade'][0] == 'comparação':
                        
                        with open('Nome_txt.txt', 'r') as f:
                            Nome_txt = f.readlines()
                            Nome_txt = [line.strip() for line in Nome_txt]
                        mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")                
                        mail.Send()     
    
                    else:  
                        if nome_loja == 'Corporativo':
                            pass 
                        else:
                            self.atualizacao.append(f"A unidade {nome_loja}_fatura não possui cadastro no banco de dados")
        
        
                    #Remove o arquivo temporario que criei
                    os.remove(output_path)
                        
       
       
       
        # Chama as funções 
        send_mail_aereo()
        send_mail_hotel()
        send_mail_fatura()
        self.atualizacao.append('E-mails enviado com sucesso!!')
    
    
        # Dá planilha que ele cria as abas com centro de custo corporativo, 
        # para cada centro de custo enviaremos para gestores diferentes    
        def mandar_corp(): 
            
            if 'Corporativo_aereo' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                #Ler somente a aba aereo do corporativo
                excel_corp_aereo = pd.read_excel("dados_selecionados.xlsx", sheet_name='Corporativo_aereo')

            if 'Corporativo_hotel' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                #Ler somente a aba hotel do corporativo
                excel_corp_hotel = pd.read_excel("dados_selecionados.xlsx", sheet_name='Corporativo_hotel')

            if 'Corporativo_fatura' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                #Ler somente a aba fatura do corporativo
                excel_corp_fatura = pd.read_excel("dados_selecionados.xlsx", sheet_name='Corporativo_fatura')


            #unificas as abas que ele leu
            with pd.ExcelWriter('dados_corporativo.xlsx') as writer:
                # Gravar cada aba em um arquivo separado no novo arquivo Excel
                if 'Corporativo_aereo' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names: 
                    excel_corp_aereo.to_excel(writer, sheet_name='Corporativo_aereo', index=False)
                if 'Corporativo_hotel' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                    excel_corp_hotel.to_excel(writer, sheet_name='Corporativo_hotel', index=False)
                if 'Corporativo_fatura' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                    excel_corp_fatura.to_excel(writer, sheet_name='Corporativo_fatura', index=False) 


            if 'Corporativo_aereo' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                # Depois de unificar vai ler cada aba 
                banco_dados_aereo_corp = pd.read_excel("dados_corporativo.xlsx", sheet_name='Corporativo_aereo')
              
            if 'Corporativo_hotel' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:        
                banco_dados_hotel_corp = pd.read_excel("dados_corporativo.xlsx", sheet_name='Corporativo_hotel')

            if 'Corporativo_fatura' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                banco_dados_fatura_corp = pd.read_excel("dados_corporativo.xlsx", sheet_name='Corporativo_fatura')


            if 'Corporativo_aereo' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                #renomear as colunas e coloca sempre o titulo com a primeira letra maiscula
                banco_dados_aereo_corp ['Centro de custo nome'] = banco_dados_aereo_corp['Centro de custo'].str.title()

            if 'Corporativo_hotel' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                #renomear as colunas  e coloca sempre o titulo com a primeira letra maiscula
                banco_dados_hotel_corp['Centro de custo nome'] = banco_dados_hotel_corp['Centro de custo'].str.title()

            if 'Corporativo_fatura' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                #renomear as colunas e coloca sempre o titulo com a primeira letra maiscula
                banco_dados_fatura_corp['Centro de custo nome'] = banco_dados_fatura_corp['CENTRO CUSTO'].str.title()

            #Le o arquivo em excel e e coloca sempre o titulo com a primeira letra maiscula
            workbook = load_workbook("dados_corporativo.xlsx")
            
            grupos_corp = banco_dados_aereo_corp.groupby('Centro de custo nome')

            grupos_hotel_corp = banco_dados_hotel_corp.groupby('Centro de custo nome')   

            if 'Corporativo_fatura' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                grupos_fatura_corp = banco_dados_fatura_corp.groupby('Centro de custo nome')
                

            # Ele ve quais os centros de de custos iguais e cria uma nova aba chamada com o centro de custo e coloca 'Aereo' no nome na aba AEREO 
            def aba_aereo_corp():
               
                if 'Corporativo_aereo' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                    for nome_loja, grupo_corp in grupos_corp:
                        dados_loja_aereo = grupo_corp   
                         
                        if not dados_loja_aereo.empty:
                            worksheet = workbook.create_sheet(f"{nome_loja}_aereo")
                            worksheet.append(['Nome das Colunas'])
                            for index, row in dados_loja_aereo.iterrows():
                                worksheet.append(row.tolist())
                            workbook.save(filename='dados_corporativo.xlsx')
                        else:
                            print(f'Sem gastos aereos na unidade {nome_loja}')

                        #conecta com o otlook
                        outlook = win32.Dispatch('outlook.application')

                        # Cria um novo e-mail
                        mail = outlook.CreateItem(0)  

                        # ler o excel 
                        excel_path = 'dados_corporativo.xlsx'

                        # abas que quero com o nome das lojas criadas no código acima
                        aba_desejada = [f"{nome_loja}_aereo"]

                        # Define o assunto
                        mail.Subject = f'Despesas aereas no Centro de Custo {aba_desejada[0]}, numero da fatura {numero_aereo}'.replace("_aereo", "")

                        # Define o corpor do e-mail
                        mail.Body = '[NAO RESPONDER ESSE EMAIL] \nOlá, encaminho em anexo os gastos de viagens lançados em seu centro de custo. '
                
                        df_dict = {}

                        # Se tiver as abas selecionadas
                        for sheet_name in aba_desejada:
                            try:
                                df_dict[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)
                            except:
                                print(f"A aba {sheet_name} não existe no arquivo Excel")
                                
                        # salva as abas em um novo arquivo
                        output_path = f'{nome_loja}.xlsx'

                        with pd.ExcelWriter(output_path) as writer:
                            for sheet_name, df in df_dict.items():
                                df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Anexa o arquivo Excel temporário (que contém apenas a aba correspondente)
                        attachment = os.path.abspath(output_path)
                        mail.Attachments.Add(attachment)
                        
                        excel = pd.read_excel(f'{nome_loja}.xlsx')
                        
                        
                        # Centro de custo Juridico
                        if excel['Centro de custo nome'][0] == 'comparacao':    
                            with open('email_653000_corporativo.txt', 'r') as f:
                                email_653000_corporativo = f.readlines()
                                email_653000_corporativo = [line.strip() for line in email_653000_corporativo]
                            mail.To = f'{email_653000_corporativo}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                            mail.Send()
                    
                        #centro de custo Orçamento / Folha / Analytics
                        elif excel['Centro de custo nome'][0] in ['comparação' or 'comparação']:    
                            with open('Nome_txt.txt', 'r') as f:
                                Nome_txt = f.readlines()
                                Nome_txt = [line.strip() for line in Nome_txt]
                            mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                            mail.Send()

                        #Centro de custo Gestao
                        elif excel['Centro de custo nome'][0] == 'comparação':           
                            with open('Nome_txt.txt', 'r') as f:
                                Nome_txt = f.readlines()
                                Nome_txt = [line.strip() for line in Nome_txt]
                            mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                            mail.Send()
                            
                        else:
                            self.atualizacao.append(f"O centro de custo {nome_loja}_hotel não possui cadastro no banco de dados")

                        # exclui o excel que ele criou temporario    
                        # os.remove(output_path)  
        
            def aba_hotel_corp():
                if 'Corporativo_hotel' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                    for nome_loja, grupo_hotel_corp in grupos_hotel_corp:
                        # selecionar os dados relevantes para cada loja
                        dados_loja_hotel = grupo_hotel_corp
                        
                        # Ele ve quais os centros de de custos iguais e cria uma nova aba chamada com o centro de custo e coloca 'Hotel' no nome na aba HOTEL         
                        if not dados_loja_hotel.empty:
                            worksheet = workbook.create_sheet(f"{nome_loja}_hotel")
                            worksheet.append(['Nome das colunas' ])
                            for index, row in dados_loja_hotel.iterrows():
                                worksheet.append(row.tolist())
                            workbook.save(filename='dados_corporativo.xlsx')
                        else: 
                            print(f'Sem gastos hoteleiros na unidade {nome_loja}')
                        
                        #conecta com o otlook
                        outlook = win32.Dispatch('outlook.application')

                        # Cria um novo e-mail
                        mail = outlook.CreateItem(0)  

                        # ler o excel 
                        excel_path = 'dados_corporativo.xlsx'

                        # abas que quero com o nome das lojas criadas no código acima
                        aba_desejada = [f"{nome_loja}_hotel"]

                        # Define o assunto
                        mail.Subject = f'Despesas com hoteis no Centro de Custo {aba_desejada[0]}, numero da fatura {numero_hotel}'.replace("_hotel", "")

                        # Define o corpor do e-mail
                        mail.Body = '[NAO RESPONDER ESSE EMAIL] \nOlá, encaminho em anexo os gastos de viagens lançados em seu centro de custo.'
                
                        df_dict = {}

                        # Se tiver as abas selecionadas
                        for sheet_name in aba_desejada:
                            try:
                                df_dict[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)
                            except:
                                print(f"A aba {sheet_name} não existe no arquivo Excel")
                                
                        # salva as abas em um novo arquivo
                        output_path = f'{nome_loja}.xlsx'

                        with pd.ExcelWriter(output_path) as writer:
                            for sheet_name, df in df_dict.items():
                                df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Anexa o arquivo Excel temporário (que contém apenas a aba correspondente)
                        attachment = os.path.abspath(output_path)
                        mail.Attachments.Add(attachment)
                        
                        excel = pd.read_excel(f'{nome_loja}.xlsx')
                        
                        
                        
                        # Centro de custo Juridico
                        if excel['Centro de custo nome'][0] == 'comparação':        
                            with open('Nome_txt.txt', 'r') as f:
                                Nome_txt = f.readlines()
                                Nome_txt = [line.strip() for line in Nome_txt]
                            mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                            mail.Send()
                    
                        #centro de custo Orçamento / Folha / Analytics
                        elif excel['Centro de custo nome'][0] in ['comparação' or 'comparação']:         
                            with open('Nome_txt.txt', 'r') as f:
                                Nome_txt = f.readlines()
                                Nome_txt = [line.strip() for line in Nome_txt]
                            mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                            mail.Send()

                           
                            
                        else:
                            self.atualizacao.append(f"O centro de custo {nome_loja}_hotel não possui cadastro no banco de dados")
            
                            
                        # exclui o excel que ele criou temporario    
                        os.remove(output_path)            
                
            # Ele ve quais os centros de de custos iguais e cria uma nova aba chamada com o centro de custo e coloca 'Fatura' no nome na aba FATURA    
            def aba_fatura_corp():
                
                
                if 'Corporativo_fatura' in pd.ExcelFile("dados_selecionados.xlsx").sheet_names:
                    for nome_loja, grupo_fatura_corp in grupos_fatura_corp:
                    
                        dados_loja_fatura = grupo_fatura_corp

                        if not dados_loja_fatura.empty:
                            worksheet = workbook.create_sheet(f"{nome_loja}_fatura")
                            worksheet.append(['Nome das Colunas'])
                            for index, row in dados_loja_fatura.iterrows():
                                worksheet.append(row.tolist())
                            workbook.save(filename='dados_corporativo.xlsx') 
                        else: 
                            print(f'Sem gastos na fatura da unidade {nome_loja}')
                        
                        
                        #conecta com o otlook
                        outlook = win32.Dispatch('outlook.application')

                        # Cria um novo e-mail
                        mail = outlook.CreateItem(0)  

                        # ler o excel 
                        excel_path = 'dados_corporativo.xlsx'

                        # abas que quero com o nome das lojas criadas no código acima
                        aba_desejada = [f"{nome_loja}_fatura"]

                        # Define o assunto
                        mail.Subject = f'Despesas do Centro de Custo {aba_desejada[0]}'.replace("_fatura", "")

                        # Define o corpor do e-mail
                        mail.Body = '[NAO RESPONDER ESSE EMAIL] \nOlá, encaminho em anexo os gastos de viagens lançados em seu centro de custo.'
                
                        df_dict = {}

                        # Se tiver as abas selecionadas
                        for sheet_name in aba_desejada:
                            try:
                                df_dict[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)
                            except:
                                print(f"A aba {sheet_name} não existe no arquivo Excel")
                                
                        # salva as abas em um novo arquivo
                        output_path = f'{nome_loja}.xlsx'

                        with pd.ExcelWriter(output_path) as writer:
                            for sheet_name, df in df_dict.items():
                                df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Anexa o arquivo Excel temporário (que contém apenas a aba correspondente)
                        attachment = os.path.abspath(output_path)
                        mail.Attachments.Add(attachment)
                        
                        excel = pd.read_excel(f'{nome_loja}.xlsx')
                        
                        
                        
                        # Centro de custo Juridico
                        if excel['Centro de custo nome'][0] == 'comparação':        
                            with open('Nome_txt.txt', 'r') as f:
                                Nome_txt = f.readlines()
                                Nome_txt = [line.strip() for line in Nome_txt ]
                            mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                            mail.Send()
                    
                        #centro de custo Orçamento / Folha / Analytics
                        elif excel['Centro de custo nome'][0] in ['comparação' or 'comparação']:        
                            with open('Nome_txt.txt', 'r') as f:
                                Nome_txt = f.readlines()
                                Nome_txt = [line.strip() for line in Nome_txt]
                            mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                            mail.Send()

                        #Centro de custo Gestao
                        elif excel['Centro de custo nome'][0] == 'comparação':         
                            with open('Nome_txt.txt', 'r') as f:
                                Nome_txt = f.readlines()
                                Nome_txt = [line.strip() for line in Nome_txt]
                            mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                            mail.Send()
                            
                        
                            
                        else: 
                            self.atualizacao.append(f"O centro de custo {nome_loja}_fatura não possui cadastro no banco de dados")

                            
                        # exclui o excel que ele criou temporario    
                        os.remove(output_path)            
            
               
        
            aba_aereo_corp()
            aba_hotel_corp()
            aba_fatura_corp()
        
        
        mandar_corp()
        
        # Se enviar os e-mail vai preencher o estatus que já enviou o e-mail        
        self.atualizacao.append('Todos os e-mails do corporativo enviado com sucesso!!')

    
       
    def consultar_unidade(self):
         
        
        # no campo que o usuario colocou a ação vai puxar as informações
        acao = self.escolha_acao.currentText()
        # no campo ele vai colocar o e-mail que queira enviar 
        email = self.text_email.text()
    
        

        # vai colocar a unidade que quiser para fazer a ação 
        unidade_selecionada = self.escolha_unidade.currentText()

        # Vai retirar e torcar caracteres da unidade que o usario quis 
        unidade = f'{str(unidade_selecionada)}'.replace('edit_', '')
        texto = f'{str(unidade_selecionada)}'.replace('edit_', 'email_')
        
        # função para criar / editar / ler o banco de dados
        def edit_corporativo():
            
            # vai consultar se aquela unidade já possui no banco de dados, se não possui vai criar um com o nome daquela unidade
            if os.path.exists(f'email_{unidade}.txt'):
                with open(f'email_{unidade}.txt', "r") as f:
                    texto = f.read().splitlines()
            else: 
                with open(f'email_{unidade}.txt', "w") as f:
                    f.write(f"{texto}.txt")
                texto = []
                
            # Se a ação for adicionar o e-mail, o usuario vai adicionar o e-mail e o comando vai colocar no banco de dados    
            if f'{str(acao)}' == 'adicionar':
                # Adiciona um novo email à lista
                texto.append(f'{email}')
            # Remove um email existente da lista
            if f'{str(acao)}' == 'remover':
                if email in texto:
                    texto.remove(f'{email}')

            # Salva a lista no arquivo
            with open(f'email_{unidade}.txt', "w") as f:
                f.write('\n'.join(texto))

        
        # pega a unidade selecionada     
        txt = f'{str(unidade_selecionada)}.txt'.replace('edit_', 'email_')
        
        # a ção 
        if str(acao) == 'Ver e-mails cadastrados':
            with open(txt, 'r') as f:
                txt = f.readlines()
            txt = [line.strip() for line in txt]    
            self.text_consult.setText(f'{txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";"))
        
        # exec(str(unidade_selecionada)+'()')
        edit_corporativo()
        
   
    def send_email_outros(self):
       
    
        # caminho da planilha da  que o usuario coloca 
        caminho_excel_1 = self.planilha_1.text()    
         
        # caminho da planilha da  que o usuario coloca  
        caminho_excel_1 = self.planilha_2.text()
         
        # caminho da planilha da  que o usuario coloca   
        caminho_excel_3 = self.planilha_3.text()
            
        
        
        # caminho o if é caso o usuario não coloque a planilha 
        if caminho_excel_1:
            caminho_excel_1 = pd.read_excel(f'{caminho_excel_1}'.replace('"',''))
        else: 
            caminho_excel_1 = None

        # caminho o if é caso o usuario não coloque a planilha 
        if caminho_excel_2:
            caminho_excel_2 = pd.read_excel(f'{caminho_excel_2}'.replace('"',''))
        else: 
            caminho_excel_2 = None
            
        # caminho o if é caso o usuario não coloque a planilha     
        if caminho_excel_3:
            # caminho 
            caminho_excel_3 = pd.read_excel(f'{caminho_excel_3}'.replace('"',''))
        else:
            caminho_excel_3 = None
        
        
        # Pega a planilha e seleciona as colunas desejadas
        colunas_1 = ['CLIENTE', 'SOLICITANTE', 'PASSAGEIRO','CENTRO CUSTO', 'APROVADOR', 'OS', 'SERVIÇO', 'FATURA', 'FEE', 'TOTAL']
        
        colunas_2 = ['CLIENTE', 'SOLICITANTE', 'PASSAGEIRO','CENTRO CUSTO', 'APROVADOR', 'OS', 'SERVIÇO', 'FATURA', 'FEE', 'TOTAL']

        coluna_3 = ['CLIENTE', 'SOLICITANTE', 'PASSAGEIRO','CENTRO CUSTO', 'APROVADOR', 'OS', 'SERVIÇO', 'FATURA', 'FEE', 'TOTAL']    
       
       
        
        # unifica as colunas que quero com a planilha 
        if caminho_excel_1 is not None:
            dados_1 = caminho_excel_1[colunas_1]
        else :
            dados_1 = None
        

        if caminho_excel_2 is not None:
            dados_2 = caminho_excel_2[colunas_2]
        else: 
            dados_2 = None
                
                
        if caminho_excel_3 is not None:
            dados_3 = caminho_excel_3[coluna_3]
        else: 
            dados_3 = None

       
       #Cria uma nova planilha somente se for passado o caminho da planilha 
        with pd.ExcelWriter('dados_outros.xlsx', engine='openpyxl') as writer:
            if caminho_excel_1 is not None:
                dados_1.to_excel(writer, sheet_name='1', index=False)
            
            if caminho_excel_2 is not None:
                dados_2.to_excel(writer, sheet_name='2', index=False)
                
            if caminho_excel_3 is not None:    
                dados_3.to_excel(writer, sheet_name='3', index=False)       
       
       
        # Ler a planilha 'aereo' no arquivo Excel
        if caminho_excel_1 is not None:
            banco_dados_1 = pd.read_excel("dados_outros.xlsx", sheet_name='1')
        
        if caminho_excel_2 is not None:
            banco_dados_2 = pd.read_excel("dados_outros.xlsx", sheet_name='2')
        
        if caminho_excel_3 is not None:   
            banco_dados_3 = pd.read_excel("dados_outros.xlsx", sheet_name='3')
       
       
        if caminho_excel_1 is not None:
            # Separa o centro de custo 1
            banco_dados_1['Centro de custo nome'] = banco_dados_1['CENTRO CUSTO']
        
        if caminho_excel_2 is not None:
            # Separa o centro de custo 2
            banco_dados_2['Centro de custo nome'] = banco_dados_2['CENTRO CUSTO']
        
        if caminho_excel_3 is not None:
            #Sepra o centro de custo 3
            banco_dados_3['Centro de custo nome'] = banco_dados_3['CENTRO CUSTO']
       
       
        # carregar o arquivo Excel existente
        workbook = load_workbook("dados_outros.xlsx")
       
       
        self.atualizacao_2.setText('-=-=-Informacões!-==-=-')
       
       
        # agrupar os dados por 'Centro de custo nome'
        if caminho_excel_1 is not None:
            grupos = banco_dados_1.groupby('Centro de custo nome')
       
        if caminho_excel_2 is not None:
            grupos = banco_dados_2.groupby('Centro de custo nome')
               
        if caminho_excel_3 is not None:       
            grupos = banco_dados_3.groupby('Centro de custo nome')
           

        # Vai separar a planilha da telmac nos centros de custos e criar abas
        def aba_1():
            if caminho_excel_1 is not None:
                # selecionar os dados relevantes para cada loja
               

                for nome_loja, grupo in grupos:
                    
                    dados_loja_1 = grupo
                     
                    if caminho_excel_1 is not None:    
                        if not dados_loja_1.empty:
                            worksheet = workbook.create_sheet(nome_loja)
                            worksheet.append(['CLIENTE', 'SOLICITANTE', 'PASSAGEIRO','CENTRO CUSTO', 'APROVADOR', 'OS', 'SERVIÇO', 'FATURA', 'FEE', 'TOTAL', 'Centro de Custo'])
                            for index, row in dados_loja_1.iterrows():
                                worksheet.append(row.tolist())
                            workbook.save(filename='dados_outros.xlsx')
                        else:
                            print(f'Sem gastos aereos na unidade {nome_loja}')

                    #conectar com outlook
                    outlook = win32.Dispatch('outlook.application')

                    # Cria um novo e-mail
                    mail = outlook.CreateItem(0)  

                    # ler o excel 
                    excel_path = 'dados_outros.xlsx'

                    # abas que quero 
                    aba_desejada = [f"{nome_loja}"]

                    # Define o assunto
                    mail.Subject = f'Custo de despesas da unidade de {aba_desejada[0]}'

                    mail.Body = '[NAO RESPONDER ESSE EMAIL] \nOlá, encaminho em anexo os gastos de viagens lançados em seu centro de custo. '

                    df_dict = {}        

                    # Se tiver as abas selecionadas
                    for sheet_name in aba_desejada:
                        try:
                            df_dict[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)
                        except:
                            print(f"A aba {sheet_name} não existe no arquivo Excel")
                            
                            
                            
                    # salva as abas em um novo arquivo
                    output_path = f'{nome_loja}.xlsx'

                    with pd.ExcelWriter(output_path) as writer:
                        for sheet_name, df in df_dict.items():
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        
                    # Anexa o arquivo Excel temporário (que contém apenas a aba correspondente)
                    attachment = os.path.abspath(output_path)
                    mail.Attachments.Add(attachment)
                    
                    excel = pd.read_excel(f'{nome_loja}.xlsx')
                
                    
                    #Compara se realmente no centro de custo tem a telmac
                    if excel['CENTRO CUSTO'][0] == 'comparacao':        
                        with open('Nome_txt.txt', 'r') as f:
                            Nome_txt = f.readlines()
                            Nome_txt = [line.strip() for line in Nome_txt]
                        mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                        mail.Send()

                    #Compara se realmente no centro de custo tem a telmac
                    elif excel['CENTRO CUSTO'][0] == 'comparacao':        
                        with open('Nome_txt.txt', 'r') as f:
                            Nome_txt = f.readlines()
                            Nome_txt = [line.strip() for line in Nome_txt]
                        mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                        mail.Send()

                    elif excel['CENTRO CUSTO'][0] == 'comparacao':
                        with open('Nome_txt.txt', 'r') as f:
                            Nome_txt = f.readlines()
                            Nome_txt = [line.strip() for line in Nome_txt]
                        mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                        mail.Send()


        
        # Vai separar a planilha da nos centros de custos e criar abas 
        def aba_2():
            if caminho_excel_2 is not None:
                for nome_loja, grupo in grupos:
                    
                    dados_loja_2 = grupo
                    
                    if caminho_excel_2 is not None:            
                        if not dados_loja_2.empty:
                            worksheet = workbook.create_sheet(nome_loja)
                            worksheet.append(['CLIENTE', 'SOLICITANTE', 'PASSAGEIRO','CENTRO CUSTO', 'APROVADOR', 'OS', 'SERVIÇO', 'FATURA', 'FEE', 'TOTAL', 'Centro de Custo'])
                            for index, row in dados_loja_2.iterrows():
                                worksheet.append(row.tolist())
                            workbook.save(filename='dados_outros.xlsx')
                        else: 
                            print(f'Sem gastos hoteleiros na unidade {nome_loja}')
                    
                    #conectar com outlook
                    outlook = win32.Dispatch('outlook.application')

                    # Cria um novo e-mail
                    mail = outlook.CreateItem(0)  

                    # ler o excel 
                    excel_path = 'dados_outros.xlsx'

                    # abas que quero 
                    aba_desejada = [f"{nome_loja}"]

                    # Define o assunto
                    mail.Subject = f'Custo de despesas da unidade de {aba_desejada[0]}'.replace("_telmac", "")

                    mail.Body = '[NAO RESPONDER ESSE EMAIL] \nOlá, encaminho em anexo os gastos de viagens lançados em seu centro de custo. '

                    df_dict = {}        

                    # Se tiver as abas selecionadas
                    for sheet_name in aba_desejada:
                        try:
                            df_dict[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)
                        except:
                            print(f"A aba {sheet_name} não existe no arquivo Excel")
                            
                            
                            
                    # salva as abas em um novo arquivo
                    output_path = f'{nome_loja}.xlsx'

                    with pd.ExcelWriter(output_path) as writer:
                        for sheet_name, df in df_dict.items():
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        
                    # Anexa o arquivo Excel temporário (que contém apenas a aba correspondente)
                    attachment = os.path.abspath(output_path)
                    mail.Attachments.Add(attachment)
                    
                    excel = pd.read_excel(f'{nome_loja}.xlsx')

                     #Compara se realmente no centro de custo tem a telmac
                    if excel['CENTRO CUSTO'][0] == 'comparacao':        
                        with open('Nome_txt.txt', 'r') as f:
                            Nome_txt = f.readlines()
                            Nome_txt = [line.strip() for line in Nome_txt]
                        mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                        mail.Send()

                    elif excel['CENTRO CUSTO'][0] == 'comparação':        
                        with open('Nome_txt.txt', 'r') as f:
                            Nome_txt = f.readlines()
                            Nome_txt = [line.strip() for line in Nome_txt]
                        mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                        mail.Send()
        
        
        # Vai separar a planilha da  nos centros de custos e criar abas
        def aba_3():
            if caminho_excel_3 is not None:
                for nome_loja, grupo in caminho_excel_3:
                    
                    dados_loja_3 = grupo
                    
                    if caminho_excel_3 is not None:    
                        if not dados_loja_3.empty:
                            worksheet = workbook.create_sheet(f'{nome_loja}_fatura')
                            worksheet.append(['CLIENTE', 'SOLICITANTE', 'PASSAGEIRO','CENTRO CUSTO', 'APROVADOR', 'OS', 'SERVIÇO', 'FATURA', 'FEE', 'TOTAL', 'Centro de Custo'])
                            for index, row in dados_loja_3.iterrows():
                                worksheet.append(row.tolist())
                            workbook.save(filename='dados_outros.xlsx') 
                        else: 
                            print(f'Sem gastos na fatura da unidade {nome_loja}')
                
                    #conectar com outlook
                    outlook = win32.Dispatch('outlook.application')

                    # Cria um novo e-mail
                    mail = outlook.CreateItem(0)  

                    # ler o excel 
                    excel_path = 'dados_outros.xlsx'

                    # abas que quero 
                    aba_desejada = [f"{nome_loja}_fatura"]

                    # Define o assunto
                    mail.Subject = f'Custo de despesas da unidade de {aba_desejada[0]}'.replace("_telmac", "")

                    mail.Body = '[NAO RESPONDER ESSE EMAIL] \nOlá, encaminho em anexo os gastos de viagens lançados em seu centro de custos.'

                    df_dict = {}        

                    # Se tiver as abas selecionadas
                    for sheet_name in aba_desejada:
                        try:
                            df_dict[sheet_name] = pd.read_excel(excel_path, sheet_name=sheet_name)
                        except:
                            print(f"A aba {sheet_name} não existe no arquivo Excel")
                            
                            
                            
                    # salva as abas em um novo arquivo
                    output_path = f'{nome_loja}.xlsx'

                    with pd.ExcelWriter(output_path) as writer:
                        for sheet_name, df in df_dict.items():
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        
                    # Anexa o arquivo Excel temporário (que contém apenas a aba correspondente)
                    attachment = os.path.abspath(output_path)
                    mail.Attachments.Add(attachment)
                    
                    excel = pd.read_excel(f'{nome_loja}.xlsx')
                  
                    if excel['CENTRO CUSTO'][0] == 'comparacao':   
                        print('oi')     
                        with open('Nome_txt.txt', 'r') as f:
                            Nome_txt = f.readlines()
                            Nome_txt = [line.strip() for line in Nome_txt]
                        mail.To = f'{Nome_txt}'.replace('[', '').replace(']', '').replace("'", "").replace(",", ";")
                        mail.Send()
                

    
                    #Exclui a planilha que criou 
                    os.remove(f'{nome_loja}.xlsx')
        
        
        
        
        #Chama as funções
        aba_1()
        aba_2()
        aba_3()
            
        #atualiza os campos 
        self.atualizacao_2.append('Todas as planilhas enviadas com sucesso!')
       
       
       
       
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec() 
    sys.exit(app.exec())
